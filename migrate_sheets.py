"""
migrate_sheets.py — Import AskMiro GAS Ops DB into Postgres.

Source: /Users/mike/Downloads/AskMiro Ops DB.xlsx  (33 sheets, full GAS export)
Also:   /Users/mike/Downloads/AskMiro_Data_Template.xlsx (historical P&L data)
Target: Postgres operational tables

Run:  python migrate_sheets.py
"""

import os, sys, logging, json
from datetime import datetime, date
from decimal import Decimal

import openpyxl

# ── Setup ──────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger("migrate")

GAS_DB = os.environ.get("ASKMIRO_GAS_DB", "/Users/mike/Downloads/AskMiro Ops DB.xlsx")
TEMPLATE_DB = os.environ.get("ASKMIRO_TEMPLATE", "/Users/mike/Downloads/AskMiro_Data_Template.xlsx")

sys.path.insert(0, os.path.dirname(__file__))
import db_pg


# ── Helpers ────────────────────────────────────────────────────────────────────

def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s or s == 'None':
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:26], fmt).date()
        except ValueError:
            continue
    log.warning("Cannot parse date: %s", s)
    return None


def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _s(val, default=""):
    """Safe string."""
    if val is None:
        return default
    return str(val).strip()


def read_sheet(wb, name):
    """Read sheet → list of dicts, skipping empty rows."""
    if name not in wb.sheetnames:
        log.warning("Sheet '%s' not found — skipping", name)
        return []
    ws = wb[name]
    if ws.max_row < 2:
        return []
    headers = [cell.value for cell in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [cell.value for cell in ws[r]]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


# ── Importers (GAS DB) ────────────────────────────────────────────────────────

def import_cleaners(conn, wb):
    """Import Cleaners sheet → ops_cleaners."""
    rows = read_sheet(wb, "Cleaners")
    imported = 0
    for r in rows:
        name = _s(r.get("fullName"))
        if not name:
            continue
        email = _s(r.get("email"))

        # Check duplicate by email
        if email:
            existing = db_pg.fetchval(conn, "SELECT id FROM ops_cleaners WHERE email = %s", (email,))
            if existing:
                log.info("  Skip dup cleaner: %s (%s)", name, email)
                continue

        cleaner_type = _s(r.get("cleanerType"), "Employee")
        if cleaner_type not in ("Employee", "Subcontractor", "Agency", "Trial"):
            cleaner_type = "Subcontractor" if "sub" in cleaner_type.lower() else "Employee"

        status = _s(r.get("status"), "Active")
        if status not in ("Active", "Inactive", "Archived", "Trial"):
            status = "Active"

        compliance = _s(r.get("complianceStatus"), "Pending")
        if compliance not in ("Ready", "Pending", "Expiring", "Blocked"):
            compliance = "Pending"

        dbs = _s(r.get("dbsStatus"), "None")
        if dbs not in ("Enhanced", "Basic", "None", "Expired"):
            dbs = "None"

        db_pg.execute(conn, """
            INSERT INTO ops_cleaners (
                full_name, email, phone, home_postcode, borough,
                cleaner_type, status, services_offered,
                availability_type, currently_available,
                compliance_status, dbs_status, transport_mode,
                hourly_rate, emergency_cover, notes
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            name, email,
            _s(r.get("phone")).replace(".0", ""),
            _s(r.get("homePostcode")),
            _s(r.get("borough")),
            cleaner_type, status,
            _s(r.get("servicesOffered")),
            _s(r.get("availabilityType"), "Full-time"),
            _s(r.get("currentlyAvailable"), "Yes"),
            compliance, dbs,
            _s(r.get("transportMode"), "Public Transport"),
            _safe_float(r.get("hourlyRate"), 12.50),
            _s(r.get("emergencyCover"), "No"),
            _s(r.get("notes")),
        ))
        imported += 1
        log.info("  Cleaner: %s", name)

    log.info("Cleaners imported: %d", imported)
    return imported


def import_labour_workers(conn, wb):
    """Import Labour_Workers → pay_workers."""
    rows = read_sheet(wb, "Labour_Workers")
    imported = 0
    for r in rows:
        name = _s(r.get("name"))
        if not name:
            continue
        email = _s(r.get("email"))

        existing = db_pg.fetchval(conn, "SELECT id FROM pay_workers WHERE full_name = %s", (name,))
        if existing:
            log.info("  Skip dup worker: %s", name)
            continue

        role = _s(r.get("role"), "Cleaner")
        if role not in ("Cleaner", "Supervisor", "Team Leader", "Driver", "Office"):
            role = "Office" if "director" in role.lower() or "manager" in role.lower() else "Cleaner"

        payment = _s(r.get("payment_method"), "BACS")
        if payment not in ("BACS", "Cheque", "Cash"):
            payment = "BACS"

        db_pg.execute(conn, """
            INSERT INTO pay_workers (
                full_name, role, default_hourly_rate, phone, email, status,
                ni_number, tax_code, payment_method
            ) VALUES (%s,%s,%s,%s,%s,'active',%s,%s,%s)
        """, (
            name, role,
            _safe_float(r.get("default_hourly_rate"), 13.85),
            _s(r.get("phone")),
            email,
            _s(r.get("ni_number")),
            _s(r.get("tax_code"), "1257L"),
            payment,
        ))
        imported += 1
        log.info("  Worker: %s (%s)", name, role)

    log.info("Workers imported: %d", imported)
    return imported


def import_labour_entries(conn, wb):
    """Import Labour_Entries → pay_entries."""
    rows = read_sheet(wb, "Labour_Entries")
    imported = 0
    for r in rows:
        worker_name = _s(r.get("worker_name"))
        if not worker_name:
            continue

        work_date = _parse_date(r.get("date"))
        if not work_date:
            continue

        # Find worker_id
        worker_id = db_pg.fetchval(conn,
            "SELECT id FROM pay_workers WHERE full_name = %s", (worker_name,))

        hours = _safe_float(r.get("hours_worked"))
        rate = _safe_float(r.get("hourly_rate"))
        total = _safe_float(r.get("total_pay"), hours * rate)

        entry_type = _s(r.get("entry_type"), "Basic Hours")
        if entry_type not in ("Basic Hours", "Overtime x1.5", "Overtime x2", "Night Shift", "Holiday Pay", "Training", "Other"):
            entry_type = "Basic Hours"

        status = _s(r.get("status"), "pending")
        if status not in ("pending", "approved", "paid"):
            status = "pending"

        db_pg.execute(conn, """
            INSERT INTO pay_entries (
                worker_id, worker_name, entry_date, hours_worked,
                hourly_rate, total_pay, entry_type, status, notes
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            worker_id, worker_name, work_date,
            hours, rate, total, entry_type, status,
            _s(r.get("notes")),
        ))
        imported += 1

    log.info("Labour entries imported: %d", imported)
    return imported


def import_finance_expenses(conn, wb):
    """Import Finance_Expenses → fin_expenses."""
    rows = read_sheet(wb, "Finance_Expenses")
    imported = 0
    for r in rows:
        expense_date = _parse_date(r.get("expenseDate"))
        if not expense_date:
            continue

        category = _s(r.get("category"), "Miscellaneous")
        db_pg.execute(conn, """
            INSERT INTO fin_expenses (
                expense_date, category, subcategory, supplier, description,
                amount_net, amount_vat, amount_gross, site_id, recurring
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            expense_date, category,
            _s(r.get("subcategory")),
            _s(r.get("supplier")),
            _s(r.get("description")),
            _safe_float(r.get("amountNet")),
            _safe_float(r.get("amountVat")),
            _safe_float(r.get("amountGross")),
            _s(r.get("linkedSiteId")) or None,
            "Yes" if _s(r.get("recurringFlag")) == "Yes" else "No",
        ))
        imported += 1

    log.info("Finance expenses imported: %d", imported)
    return imported


def import_finance_transactions(conn, wb):
    """Import Finance_Transactions → fin_transactions."""
    rows = read_sheet(wb, "Finance_Transactions")
    imported = 0
    for r in rows:
        tx_date = _parse_date(r.get("transactionDate"))
        if not tx_date:
            continue

        tx_type = _s(r.get("type"), "expense")
        if tx_type not in ("income", "expense"):
            tx_type = "expense"

        db_pg.execute(conn, """
            INSERT INTO fin_transactions (
                transaction_date, type, category, description,
                amount_gross, external_ref, site_id, notes
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            tx_date, tx_type,
            _s(r.get("category"), "Miscellaneous"),
            _s(r.get("description")),
            _safe_float(r.get("amountGross")),
            _s(r.get("externalRef")) or _s(r.get("id")),
            _s(r.get("linkedSiteId")) or None,
            _s(r.get("notes")),
        ))
        imported += 1

    log.info("Finance transactions imported: %d", imported)
    return imported


def import_finance_settings(conn, wb):
    """Import Finance_Settings → fin_settings."""
    rows = read_sheet(wb, "Finance_Settings")
    # Map GAS keys to our fin_settings keys
    key_map = {
        "vatRate": "vat_rate",
        "targetMarginHealthy": "target_margin_healthy",
        "targetMarginWatch": "target_margin_watch",
        "invoicePrefix": "invoice_prefix",
        "defaultPaymentTerms": "default_payment_terms",
        "currency": "currency",
    }
    imported = 0
    for r in rows:
        gas_key = _s(r.get("key"))
        pg_key = key_map.get(gas_key, gas_key)
        val = str(r.get("value", "")) if r.get("value") is not None else ""
        # Clean float strings
        if val.endswith(".0"):
            val = val[:-2]

        db_pg.execute(conn, """
            INSERT INTO fin_settings (key, value) VALUES (%s, %s)
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """, (pg_key, val))
        imported += 1

    log.info("Finance settings imported: %d", imported)
    return imported


def import_leads_from_gas(conn, wb):
    """LEGACY — kept for reference. Use import_leads_normalized instead."""
    log.warning("import_leads_from_gas is DEPRECATED — use import_leads_normalized")
    return 0


def import_outreach_log(conn, wb):
    """LEGACY — kept for reference. Use import_outreach_normalized instead."""
    log.warning("import_outreach_log is DEPRECATED — use import_outreach_normalized")
    return 0


# ── Normalized Importers (replace legacy fin_transactions hacks) ─────────────

def _parse_timestamp(val):
    """Parse a value into a datetime (not just date). Returns datetime or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    if not s or s == 'None':
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:26], fmt)
        except ValueError:
            continue
    log.warning("Cannot parse timestamp: %s", s)
    return None


def _clean_phone(val):
    """Clean phone string: remove .0 suffix, ensure string."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _map_stage(status):
    """Map GAS lead status → pipeline_stage ENUM value."""
    mapping = {
        'contacted': 'contacted',
        'outreachqueued': 'ready_to_contact',
        'new': 'new',
        'qualified': 'replied',
        'won': 'won',
        'lost': 'lost',
        'replied': 'replied',
        'meeting': 'meeting_or_site_visit',
        'quotesent': 'quote_sent',
        'negotiating': 'negotiating',
        'dormant': 'dormant',
        'enriched': 'enriched',
    }
    return mapping.get(status.lower().strip(), 'new') if status else 'new'


def _score_band(score):
    """Return A/B/C/D band from total score."""
    if score >= 80:
        return 'A'
    if score >= 60:
        return 'B'
    if score >= 40:
        return 'C'
    return 'D'


def _next_best_action(status):
    """Derive next best action from lead status."""
    mapping = {
        'Contacted': 'Follow up — awaiting reply',
        'OutreachQueued': 'Send initial outreach email',
        'New': 'Enrich and score lead',
        'Qualified': 'Schedule site visit or call',
        'Won': 'Onboard client',
        'Lost': 'Re-engage in 90 days',
    }
    return mapping.get(status, 'Review and qualify')


def import_leads_normalized(conn, wb):
    """Import 372 GAS Leads → normalized schema: entities, addresses,
    entity_locations, contacts, opportunities, opportunity_scores, entity_source_links.
    """
    rows = read_sheet(wb, "Leads")
    imported = 0
    skipped = 0

    for r in rows:
        company = _s(r.get("companyName"))
        if not company:
            continue
        email = _s(r.get("email"))
        gas_id = _s(r.get("id"))  # e.g. LEAD-MN5RXQ7AH4MY
        phone = _clean_phone(r.get("phone"))
        contact_name = _s(r.get("contactName"))
        segment = _s(r.get("segment"))
        status = _s(r.get("status"), "New")
        postcode = _s(r.get("postcode"))
        notes = _s(r.get("notes"))
        lead_score = _safe_int(r.get("leadScore"), 50)
        created_at = _parse_timestamp(r.get("createdAt"))
        updated_at = _parse_timestamp(r.get("updatedAt"))
        next_followup = _parse_timestamp(r.get("nextFollowUpAt"))

        normalized_name = company.lower().strip()

        # ── Deduplicate: check entity_source_links first ──
        if gas_id:
            existing_eid = db_pg.fetchval(conn,
                "SELECT entity_id FROM entity_source_links WHERE source = 'manual'::source_system AND source_record_id = %s",
                (gas_id,))
            if existing_eid:
                skipped += 1
                continue

        # Also dedup by normalized_name (covers scraper-imported entities)
        existing_eid = db_pg.fetchval(conn,
            "SELECT id FROM entities WHERE normalized_name = %s LIMIT 1",
            (normalized_name,))
        if not existing_eid and email:
            # Try by email
            existing_eid = db_pg.fetchval(conn,
                "SELECT id FROM entities WHERE primary_email = %s LIMIT 1",
                (email,))
        if existing_eid:
            skipped += 1
            # Update email/phone if entity is missing them
            db_pg.execute(conn, """
                UPDATE entities SET
                    primary_email = COALESCE(NULLIF(primary_email, ''), %s),
                    primary_phone = COALESCE(NULLIF(primary_phone, ''), %s),
                    updated_at = NOW()
                WHERE id = %s
            """, (email or None, phone or None, existing_eid))
            # Ensure source link exists
            if gas_id:
                db_pg.execute(conn, """
                    INSERT INTO entity_source_links (entity_id, source, source_record_id, confidence)
                    VALUES (%s, %s::source_system, %s, 100)
                    ON CONFLICT (source, source_record_id) DO NOTHING
                """, (existing_eid, 'manual', gas_id))
            # Ensure opportunity exists for this entity
            opp_exists = db_pg.fetchval(conn,
                "SELECT id FROM opportunities WHERE entity_id = %s", (existing_eid,))
            if not opp_exists:
                stage = _map_stage(status)
                db_pg.execute(conn, """
                    INSERT INTO opportunities (entity_id, title, current_stage,
                                               last_touched_at, next_followup_at, notes)
                    VALUES (%s, %s, %s::pipeline_stage, %s, %s, %s)
                    ON CONFLICT (entity_id) DO NOTHING
                """, (existing_eid, f"Commercial cleaning — {company}", stage,
                      updated_at, next_followup, (notes or "")[:500] or None))
            # Ensure contact exists
            contact_exists = db_pg.fetchval(conn,
                "SELECT id FROM contacts WHERE entity_id = %s AND is_primary = TRUE", (existing_eid,))
            if not contact_exists and (contact_name or email):
                db_pg.execute(conn, """
                    INSERT INTO contacts (entity_id, full_name, email, phone, is_primary)
                    VALUES (%s, %s, %s, %s, TRUE)
                """, (existing_eid, contact_name or None, email or None, phone or None))
            imported += 1
            if imported % 50 == 0:
                log.info("  ... processed %d leads so far (matched existing)", imported)
            continue

        # ── Map sector ──
        sector_map = {
            'office': 'Office',
            'residential': 'Residential',
            'commercial': 'Commercial',
            'retail': 'Retail',
            'healthcare': 'Healthcare',
            'education': 'Education',
            'hospitality': 'Hospitality',
            'industrial': 'Industrial',
        }
        sector = sector_map.get((segment or '').lower().strip(), segment or 'Commercial')

        # ── 1. Insert entity ──
        entity_id = db_pg.fetchval(conn, """
            INSERT INTO entities (entity_kind, canonical_name, normalized_name,
                                  sector, primary_phone, primary_email, active)
            VALUES (%s::entity_kind, %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, ('company', company, normalized_name, sector, phone, email))

        # ── 2. Insert address (if postcode present) ──
        address_id = None
        if postcode:
            # Derive borough from postcode district (rough mapping)
            borough = None  # Could add postcode→borough mapping later
            address_id = db_pg.fetchval(conn, """
                INSERT INTO addresses (postcode, borough)
                VALUES (%s, %s)
                RETURNING id
            """, (postcode, borough))

            # ── 3. Link entity to address ──
            db_pg.execute(conn, """
                INSERT INTO entity_locations (entity_id, address_id, is_primary, location_type)
                VALUES (%s, %s, TRUE, 'trading')
                ON CONFLICT (entity_id, address_id) DO NOTHING
            """, (entity_id, address_id))

        # ── 4. Insert contact ──
        if contact_name or email or phone:
            db_pg.execute(conn, """
                INSERT INTO contacts (entity_id, full_name, email, phone, is_primary)
                VALUES (%s, %s, %s, %s, TRUE)
            """, (entity_id, contact_name or None, email or None, phone or None))

        # ── 5. Insert opportunity ──
        stage = _map_stage(status)
        db_pg.execute(conn, """
            INSERT INTO opportunities (entity_id, title, current_stage,
                                       last_touched_at, next_followup_at, notes)
            VALUES (%s, %s, %s::pipeline_stage, %s, %s, %s)
            ON CONFLICT (entity_id) DO NOTHING
        """, (
            entity_id,
            f"Commercial cleaning — {company}",
            stage,
            updated_at,
            next_followup,
            (notes or "")[:500] or None,
        ))

        # ── 6. Insert opportunity_scores ──
        total = max(0, min(100, lead_score))
        # Distribute score proportionally across sub-scores
        # fit=25%, facility=20%, buyer_signal=25%, contactability=15%, scale=10%, freshness=5%
        fit = round(total * 0.25)
        facility = round(total * 0.20)
        buyer_signal = round(total * 0.25)
        contactability = round(total * 0.15)
        scale = round(total * 0.10)
        freshness = round(total * 0.05)

        db_pg.execute(conn, """
            INSERT INTO opportunity_scores
                (entity_id, fit_score, facility_score, buyer_signal_score,
                 contactability_score, scale_score, freshness_score,
                 total_score, score_band, next_best_action)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (entity_id) DO NOTHING
        """, (
            entity_id, fit, facility, buyer_signal,
            contactability, scale, freshness,
            total, _score_band(total), _next_best_action(status),
        ))

        # ── 7. Insert entity_source_links ──
        if gas_id:
            db_pg.execute(conn, """
                INSERT INTO entity_source_links (entity_id, source, source_record_id, confidence)
                VALUES (%s, %s::source_system, %s, 100)
                ON CONFLICT (source, source_record_id) DO NOTHING
            """, (entity_id, 'manual', gas_id))

        imported += 1
        if imported % 50 == 0:
            log.info("  ... imported %d leads so far", imported)

    log.info("Normalized leads imported: %d (skipped %d dups)", imported, skipped)
    return imported


def import_outreach_normalized(conn, wb):
    """Import 104 Outreach_Log entries → outreach_activities table."""
    rows = read_sheet(wb, "Outreach_Log")
    # Also load leads to get outreachEmailBody
    lead_rows = read_sheet(wb, "Leads")
    lead_body_map = {}
    for lr in lead_rows:
        lid = _s(lr.get("id"))
        body = _s(lr.get("outreachEmailBody"))
        if lid and body:
            lead_body_map[lid] = body

    imported = 0
    for r in rows:
        company = _s(r.get("companyName"))
        email = _s(r.get("email"))
        if not company and not email:
            continue

        sent_at = _parse_timestamp(r.get("sentAt"))
        if not sent_at:
            continue

        lead_id = _s(r.get("leadId"))
        subject = _s(r.get("subject"))
        template = _s(r.get("templateUsed"))
        reply_status = _s(r.get("replyStatus"))
        follow_up_n = _s(r.get("followUpN"))
        status = _s(r.get("status"))
        log_id = _s(r.get("logId"))

        # Look up entity_id from entity_source_links by leadId
        entity_id = None
        if lead_id:
            entity_id = db_pg.fetchval(conn, """
                SELECT entity_id FROM entity_source_links
                WHERE source = 'manual'::source_system AND source_record_id = %s
            """, (lead_id,))

        # Fallback: look up by email in entities
        if not entity_id and email:
            entity_id = db_pg.fetchval(conn,
                "SELECT id FROM entities WHERE primary_email = %s LIMIT 1", (email,))

        # Get body from lead data
        body = lead_body_map.get(lead_id, "")

        notes_str = f"Template: {template} | Follow-up #{follow_up_n} | Status: {status}"

        # Check for duplicate by log_id in notes or by entity+subject+date
        if log_id:
            existing = db_pg.fetchval(conn,
                "SELECT id FROM outreach_activities WHERE notes LIKE %s",
                (f"%{log_id}%",))
            if existing:
                continue

        # Pack all info into notes since table is minimal schema
        full_notes = f"Subject: {subject[:200]}" if subject else ""
        if template:
            full_notes += f" | Template: {template}"
        if follow_up_n:
            full_notes += f" | Follow-up #{follow_up_n}"
        if status:
            full_notes += f" | Status: {status}"
        if log_id:
            full_notes += f" | LogId: {log_id}"
        if body:
            full_notes += f"\n\nBody: {body[:1000]}"

        db_pg.execute(conn, """
            INSERT INTO outreach_activities
                (entity_id, activity_type, outcome, notes, contacted_name, logged_at)
            VALUES (%s, 'email', %s, %s, %s, %s)
        """, (
            entity_id,
            reply_status or None,
            full_notes.strip() or None,
            _s(r.get("contactName")) or None,
            sent_at,
        ))
        imported += 1

    log.info("Outreach activities imported: %d", imported)
    return imported


def import_email_log(conn, wb):
    """Import 54 EmailLog entries → email_send_log table."""
    rows = read_sheet(wb, "EmailLog")
    if not rows:
        log.info("EmailLog sheet not found or empty — skipping")
        return 0

    # Ensure email_send_log table exists
    db_pg.execute(conn, "SAVEPOINT sp_email_log_table")
    try:
        db_pg.execute(conn, """
            CREATE TABLE IF NOT EXISTS email_send_log (
                id          SERIAL PRIMARY KEY,
                email       TEXT NOT NULL,
                place_id    TEXT,
                entity_id   INTEGER,
                status      TEXT NOT NULL DEFAULT 'pending',
                detail      TEXT,
                sent_at     TIMESTAMP DEFAULT NOW(),
                delivered_at TIMESTAMP,
                bounced_at  TIMESTAMP,
                updated_at  TIMESTAMP DEFAULT NOW()
            )
        """)
        db_pg.execute(conn, "RELEASE SAVEPOINT sp_email_log_table")
    except Exception:
        db_pg.execute(conn, "ROLLBACK TO SAVEPOINT sp_email_log_table")

    imported = 0
    status_map = {
        'sent': 'sent',
        'delivered': 'delivered',
        'bounced': 'bounced',
        'failed': 'failed',
        'blocked': 'blocked',
        'pending': 'pending',
        'suppressed': 'suppressed',
    }

    for r in rows:
        to_email = _s(r.get("to"))
        if not to_email:
            continue

        ts = _parse_timestamp(r.get("ts"))
        raw_status = _s(r.get("status"), "sent").lower()
        pg_status = status_map.get(raw_status, 'sent')
        email_type = _s(r.get("type"))
        subject = _s(r.get("subject"))
        detail = f"{email_type}: {subject}" if email_type or subject else None

        # Try to match entity by email
        entity_id = db_pg.fetchval(conn,
            "SELECT id FROM entities WHERE primary_email = %s LIMIT 1", (to_email,))

        db_pg.execute(conn, """
            INSERT INTO email_send_log (email, entity_id, status, detail, sent_at)
            VALUES (%s, %s, %s, %s, %s)
        """, (to_email, entity_id, pg_status, detail, ts))
        imported += 1

    log.info("Email log entries imported: %d", imported)
    return imported


def import_quotes(conn, wb):
    """Import Quotes sheet → fin_transactions with category 'Quote'
    (temporary storage until a proper quotes table is integrated).
    """
    rows = read_sheet(wb, "Quotes")
    if not rows:
        log.info("Quotes sheet not found or empty — skipping")
        return 0

    imported = 0
    for r in rows:
        quote_id = _s(r.get("id"))  # QUOTE-xxx
        if not quote_id:
            continue

        # Dedup by external_ref
        existing = db_pg.fetchval(conn,
            "SELECT id FROM fin_transactions WHERE external_ref = %s", (quote_id,))
        if existing:
            continue

        client = _s(r.get("clientName"))
        site_addr = _s(r.get("siteAddress"))
        segment = _s(r.get("segment"))
        annual_value = _safe_float(r.get("annualValue"))
        created = _parse_date(r.get("createdAt"))

        # Collect intel_* fields into a JSON blob
        intel_fields = {}
        for k, v in r.items():
            if k and str(k).startswith("intel_") and v is not None:
                intel_fields[str(k)] = str(v)

        description = f"{client} | {site_addr} | {segment}".strip(" |")

        db_pg.execute(conn, """
            INSERT INTO fin_transactions (
                transaction_date, type, category, description,
                amount_gross, external_ref, notes
            ) VALUES (%s, 'income', 'Quote', %s, %s, %s, %s)
        """, (
            created or date.today(),
            description,
            annual_value,
            quote_id,
            json.dumps(intel_fields) if intel_fields else None,
        ))
        imported += 1

    log.info("Quotes imported: %d", imported)
    return imported


def import_intelligence_benchmarks(conn, wb):
    """Import Intelligence_Benchmarks → intelligence_benchmarks table."""
    rows = read_sheet(wb, "Intelligence_Benchmarks")
    if not rows:
        log.info("Intelligence_Benchmarks sheet not found or empty — skipping")
        return 0

    # Create table if not exists
    db_pg.execute(conn, "SAVEPOINT sp_intel_bench")
    try:
        db_pg.execute(conn, """
            CREATE TABLE IF NOT EXISTS intelligence_benchmarks (
                id SERIAL PRIMARY KEY,
                facility_type TEXT UNIQUE NOT NULL,
                mins_per_m2 NUMERIC(6,3),
                typical_m2 NUMERIC(8,1),
                supplies_per_m2_per_month NUMERIC(6,3),
                intensity_default TEXT,
                deep_clean_multiplier NUMERIC(4,2),
                notes TEXT
            )
        """)
        db_pg.execute(conn, "RELEASE SAVEPOINT sp_intel_bench")
    except Exception:
        db_pg.execute(conn, "ROLLBACK TO SAVEPOINT sp_intel_bench")

    imported = 0
    for r in rows:
        facility_type = _s(r.get("facilityType") or r.get("facility_type"))
        if not facility_type:
            continue

        # Dedup
        existing = db_pg.fetchval(conn,
            "SELECT id FROM intelligence_benchmarks WHERE facility_type = %s", (facility_type,))
        if existing:
            continue

        db_pg.execute(conn, """
            INSERT INTO intelligence_benchmarks
                (facility_type, mins_per_m2, typical_m2, supplies_per_m2_per_month,
                 intensity_default, deep_clean_multiplier, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            facility_type,
            _safe_float(r.get("minsPerM2") or r.get("mins_per_m2")) or None,
            _safe_float(r.get("typicalM2") or r.get("typical_m2")) or None,
            _safe_float(r.get("suppliesPerM2PerMonth") or r.get("supplies_per_m2_per_month")) or None,
            _s(r.get("intensityDefault") or r.get("intensity_default")) or None,
            _safe_float(r.get("deepCleanMultiplier") or r.get("deep_clean_multiplier")) or None,
            _s(r.get("notes")) or None,
        ))
        imported += 1

    log.info("Intelligence benchmarks imported: %d", imported)
    return imported


def import_intelligence_settings(conn, wb):
    """Import Intelligence_Settings → intelligence_settings table."""
    rows = read_sheet(wb, "Intelligence_Settings")
    if not rows:
        log.info("Intelligence_Settings sheet not found or empty — skipping")
        return 0

    # Create table if not exists
    db_pg.execute(conn, "SAVEPOINT sp_intel_settings")
    try:
        db_pg.execute(conn, """
            CREATE TABLE IF NOT EXISTS intelligence_settings (
                id SERIAL PRIMARY KEY,
                setting TEXT UNIQUE NOT NULL,
                value TEXT NOT NULL
            )
        """)
        db_pg.execute(conn, "RELEASE SAVEPOINT sp_intel_settings")
    except Exception:
        db_pg.execute(conn, "ROLLBACK TO SAVEPOINT sp_intel_settings")

    imported = 0
    for r in rows:
        setting = _s(r.get("Setting") or r.get("setting") or r.get("key"))
        if not setting:
            continue
        val = str(r.get("Value") or r.get("value", "")) if (r.get("Value") is not None or r.get("value") is not None) else ""
        if val.endswith(".0"):
            val = val[:-2]

        db_pg.execute(conn, """
            INSERT INTO intelligence_settings (setting, value) VALUES (%s, %s)
            ON CONFLICT (setting) DO UPDATE SET value = EXCLUDED.value
        """, (setting, val))
        imported += 1

    log.info("Intelligence settings imported: %d", imported)
    return imported


def import_finance_categories(conn, wb):
    """Import Finance_Categories → fin_expense_categories table."""
    rows = read_sheet(wb, "Finance_Categories")
    if not rows:
        log.info("Finance_Categories sheet not found or empty — skipping")
        return 0

    # Create table if not exists
    db_pg.execute(conn, "SAVEPOINT sp_fin_categories")
    try:
        db_pg.execute(conn, """
            CREATE TABLE IF NOT EXISTS fin_expense_categories (
                id SERIAL PRIMARY KEY,
                category TEXT NOT NULL,
                subcategory TEXT,
                type TEXT DEFAULT 'expense',
                active BOOLEAN DEFAULT TRUE
            )
        """)
        db_pg.execute(conn, "RELEASE SAVEPOINT sp_fin_categories")
    except Exception:
        db_pg.execute(conn, "ROLLBACK TO SAVEPOINT sp_fin_categories")

    imported = 0
    for r in rows:
        category = _s(r.get("category"))
        if not category:
            continue
        subcategory = _s(r.get("subcategory")) or None
        cat_type = _s(r.get("type"), "expense")
        active = _s(r.get("active"), "TRUE").upper() in ("TRUE", "YES", "1", "Y")

        # Dedup by category + subcategory
        existing = db_pg.fetchval(conn, """
            SELECT id FROM fin_expense_categories
            WHERE category = %s AND COALESCE(subcategory, '') = COALESCE(%s, '')
        """, (category, subcategory))
        if existing:
            continue

        db_pg.execute(conn, """
            INSERT INTO fin_expense_categories (category, subcategory, type, active)
            VALUES (%s, %s, %s, %s)
        """, (category, subcategory, cat_type, active))
        imported += 1

    log.info("Finance categories imported: %d", imported)
    return imported


# ── Importers (Template DB — historical P&L) ──────────────────────────────────

def _read_template_sheet(wb, name, header_row=3):
    """Read a template sheet where headers are on row 3 (row 1-2 are labels)."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    if ws.max_row < header_row + 1:
        return []
    headers = [cell.value for cell in ws[header_row]]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [cell.value for cell in ws[r]]
        if all(v is None for v in vals):
            continue
        if vals[0] is not None and str(vals[0]).startswith("PERIOD"):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


def import_historical_pl(conn, wb_template):
    """Import HoursFinance from AskMiro_Data_Template.xlsx → invoices + expenses + snapshots."""
    if wb_template is None:
        log.warning("Template workbook not available — skipping historical P&L")
        return 0, 0

    if "💰 HoursFinance" not in wb_template.sheetnames:
        log.warning("HoursFinance sheet not found in template")
        return 0, 0

    rows = _read_template_sheet(wb_template, "💰 HoursFinance")
    inv_count = 0
    exp_count = 0

    for r in rows:
        if not r.get("Month") or not r.get("SiteID"):
            continue
        month_date = _parse_date(r["Month"])
        if not month_date:
            continue

        site_id = _s(r["SiteID"])
        site_name = _s(r.get("SiteName", site_id))
        revenue = _safe_float(r.get("RevenueGBP"))
        labour = _safe_float(r.get("LabourCostGBP"))
        supplies = _safe_float(r.get("SuppliesCostGBP"))
        contracted = _safe_float(r.get("ContractedHours"))
        actual = _safe_float(r.get("ActualHours"))
        month_str = month_date.strftime("%Y-%m")

        # Revenue → fin_invoices
        if revenue > 0:
            existing = db_pg.fetchval(conn, """
                SELECT id FROM fin_invoices
                WHERE site_id = %s AND TO_CHAR(invoice_date, 'YYYY-MM') = %s
                  AND notes NOT LIKE 'Site reference%%'
            """, (site_id, month_str))
            if not existing:
                db_pg.execute(conn, """
                    INSERT INTO fin_invoices (entity_id, site_id, customer_name,
                        invoice_date, due_date,
                        subtotal_net, vat_amount, total_gross,
                        amount_paid, balance_due, status, notes)
                    VALUES (0, %s, %s, %s, %s, %s, 0, %s, %s, 0, 'Paid', %s)
                """, (
                    site_id, site_name,
                    month_date, date(month_date.year, month_date.month, 28),
                    revenue, revenue, revenue,
                    f"Monthly cleaning — {site_name} — {month_str} | Contracted: {contracted}h | Actual: {actual}h",
                ))
                inv_count += 1

        # Labour cost
        if labour > 0:
            existing = db_pg.fetchval(conn, """
                SELECT id FROM fin_expenses WHERE site_id = %s AND category = 'Labour'
                  AND TO_CHAR(expense_date, 'YYYY-MM') = %s
            """, (site_id, month_str))
            if not existing:
                db_pg.execute(conn, """
                    INSERT INTO fin_expenses (site_id, expense_date, category,
                        description, amount_net, amount_vat, amount_gross, supplier)
                    VALUES (%s, %s, 'Labour', %s, %s, 0, %s, 'Staff Wages')
                """, (site_id, month_date, f"Labour — {site_name} — {month_str}", labour, labour))
                exp_count += 1

        # Supplies cost
        if supplies > 0:
            existing = db_pg.fetchval(conn, """
                SELECT id FROM fin_expenses WHERE site_id = %s AND category = 'Supplies & Consumables'
                  AND TO_CHAR(expense_date, 'YYYY-MM') = %s
            """, (site_id, month_str))
            if not existing:
                db_pg.execute(conn, """
                    INSERT INTO fin_expenses (site_id, expense_date, category,
                        description, amount_net, amount_vat, amount_gross, supplier)
                    VALUES (%s, %s, 'Supplies & Consumables', %s, %s, 0, %s, 'Cleaning Supplies')
                """, (site_id, month_date, f"Supplies — {site_name} — {month_str}", supplies, supplies))
                exp_count += 1

        # Snapshots
        total_cost = labour + supplies
        gross_profit = revenue - total_cost
        margin = (gross_profit / revenue * 100) if revenue > 0 else 0
        risk = "healthy" if margin >= 35 else "watch" if margin >= 20 else "risk" if margin >= 0 else "loss"

        db_pg.execute(conn, """
            INSERT INTO fin_snapshots
                (snapshot_month, site_id, invoiced_revenue,
                 labour_cost, supplies_cost, travel_cost,
                 subcontractor_cost, other_cost, total_cost,
                 gross_profit, gross_margin_pct, risk_flag)
            VALUES (%s, %s, %s, %s, %s, 0, 0, 0, %s, %s, %s, %s)
            ON CONFLICT (snapshot_month, site_id) DO UPDATE SET
                invoiced_revenue   = EXCLUDED.invoiced_revenue,
                labour_cost        = EXCLUDED.labour_cost,
                supplies_cost      = EXCLUDED.supplies_cost,
                total_cost         = EXCLUDED.total_cost,
                gross_profit       = EXCLUDED.gross_profit,
                gross_margin_pct   = EXCLUDED.gross_margin_pct,
                risk_flag          = EXCLUDED.risk_flag
        """, (month_str, site_id, revenue, labour, supplies,
              total_cost, gross_profit, margin, risk))

    log.info("Historical P&L: %d invoices, %d expenses imported", inv_count, exp_count)
    return inv_count, exp_count


def import_template_inspections(conn, wb_template):
    """Import inspections from Data_Template."""
    if wb_template is None or "🔍 Inspections" not in wb_template.sheetnames:
        return 0
    rows = _read_template_sheet(wb_template, "🔍 Inspections")
    imported = 0
    for r in rows:
        if not r.get("Date") or not r.get("SiteID"):
            continue
        d = _parse_date(r["Date"])
        if not d:
            continue
        existing = db_pg.fetchval(conn,
            "SELECT id FROM ops_inspections WHERE site_id = %s AND inspection_date = %s",
            (r["SiteID"], d))
        if existing:
            continue
        passed = _safe_int(r.get("PassedItems"))
        total = _safe_int(r.get("TotalItems"))
        note_parts = []
        if _s(r.get("Notes")):
            note_parts.append(_s(r.get("Notes")))
        if total > 0:
            note_parts.append(f"Passed: {passed}/{total}")
        db_pg.execute(conn, """
            INSERT INTO ops_inspections (site_id, inspector, inspection_date, score, notes)
            VALUES (%s,%s,%s,%s,%s)
        """, (r["SiteID"], _s(r.get("Inspector"), "Unknown"), d,
              _safe_int(r.get("AuditScorePct")),
              " | ".join(note_parts) if note_parts else None))
        imported += 1
    log.info("Template inspections imported: %d", imported)
    return imported


def import_template_incidents(conn, wb_template):
    """Import incidents from Data_Template."""
    if wb_template is None or "⚠️ Incidents" not in wb_template.sheetnames:
        return 0
    rows = _read_template_sheet(wb_template, "⚠️ Incidents")
    imported = 0
    for r in rows:
        if not r.get("Date") or not r.get("SiteID"):
            continue
        d = _parse_date(r["Date"])
        if not d:
            continue
        sev_map = {"Low": "low", "Med": "medium", "High": "high"}
        severity = sev_map.get(_s(r.get("Severity")), "medium")
        desc = _s(r.get("Description"))
        existing = db_pg.fetchval(conn,
            "SELECT id FROM ops_incidents WHERE site_id = %s AND incident_date = %s AND description = %s",
            (r["SiteID"], d, desc))
        if existing:
            continue
        inc_type = _s(r.get("Type"), "Complaint")
        if inc_type not in ("Complaint", "Near Miss", "Accident", "Reclean"):
            inc_type = "Complaint"
        inc_status = "Resolved" if _s(r.get("Status")).lower() == "resolved" else "Open"
        resolution = _s(r.get("ResolutionNotes"))
        if _s(r.get("ReportedBy")):
            resolution = f"Reported by: {_s(r.get('ReportedBy'))} | {resolution}"
        db_pg.execute(conn, """
            INSERT INTO ops_incidents (site_id, incident_date, incident_type,
                description, status, resolution)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (r["SiteID"], d, inc_type, desc, inc_status, resolution))
        imported += 1
    log.info("Template incidents imported: %d", imported)
    return imported


def import_template_contracts(conn, wb_template):
    """Import contracts from Data_Template."""
    if wb_template is None or "📄 Contracts" not in wb_template.sheetnames:
        return 0
    rows = _read_template_sheet(wb_template, "📄 Contracts")
    imported = 0
    for r in rows:
        if not r.get("SiteID") or not r.get("SiteName"):
            continue
        contract_id = _s(r.get("ContractID"))
        existing = db_pg.fetchval(conn,
            "SELECT id FROM fin_transactions WHERE external_ref = %s AND type = 'income'",
            (contract_id,))
        if existing:
            continue
        start_date = _parse_date(r.get("StartDate"))
        renewal_date = _parse_date(r.get("RenewalDate"))
        annual_value = _safe_float(r.get("AnnualValueGBP"))
        status = _s(r.get("Status"), "Active")
        db_pg.execute(conn, """
            INSERT INTO fin_transactions (transaction_date, type, category,
                description, amount_gross, external_ref, site_id, notes)
            VALUES (%s, 'income', 'Contract Revenue', %s, %s, %s, %s, %s)
        """, (
            start_date or date.today(),
            f"Annual contract — {r['SiteName']}",
            annual_value, contract_id, r["SiteID"],
            f"Status: {status} | Renewal: {renewal_date}",
        ))
        imported += 1
    log.info("Template contracts imported: %d", imported)
    return imported


# ── Main ───────────────────────────────────────────────────────────────────────

def run_migration():
    """Execute the full migration from both workbooks."""

    # Load GAS DB (required)
    if not os.path.exists(GAS_DB):
        log.error("GAS DB not found: %s", GAS_DB)
        sys.exit(1)
    log.info("Loading GAS DB: %s", GAS_DB)
    wb_gas = openpyxl.load_workbook(GAS_DB, data_only=True)
    log.info("  Sheets: %s", wb_gas.sheetnames)

    # Load Template DB (optional — has historical P&L data)
    wb_template = None
    if os.path.exists(TEMPLATE_DB):
        log.info("Loading Template DB: %s", TEMPLATE_DB)
        wb_template = openpyxl.load_workbook(TEMPLATE_DB, data_only=True)
    else:
        log.warning("Template DB not found at %s — skipping historical P&L", TEMPLATE_DB)

    with db_pg.transaction() as conn:
        # Ensure tables exist
        try:
            from ops_tables import ensure_ops_tables
            db_pg.execute(conn, "SAVEPOINT migration_tables")
            ensure_ops_tables(conn)
            db_pg.execute(conn, "RELEASE SAVEPOINT migration_tables")
        except Exception as exc:
            db_pg.execute(conn, "ROLLBACK TO SAVEPOINT migration_tables")
            log.warning("Table creation skipped: %s", exc)

        log.info("=" * 60)
        log.info("PHASE 1: GAS Operational Data")
        log.info("=" * 60)

        # 1. Cleaners (6 records — real team + applicants)
        import_cleaners(conn, wb_gas)

        # 2. Payroll workers (2 records — Mike + Romel)
        import_labour_workers(conn, wb_gas)

        # 3. Labour entries (2 records)
        import_labour_entries(conn, wb_gas)

        # 4. Finance expenses (1 record — insurance)
        import_finance_expenses(conn, wb_gas)

        # 5. Finance transactions (1 record)
        import_finance_transactions(conn, wb_gas)

        # 6. Finance settings (6 records)
        import_finance_settings(conn, wb_gas)

        # 7. GAS Leads → NORMALIZED SCHEMA (372 leads → entities, opportunities, contacts, etc.)
        import_leads_normalized(conn, wb_gas)

        # 8. Outreach log → outreach_activities (104 email sends, properly linked)
        import_outreach_normalized(conn, wb_gas)

        # 9. Email log (54 email audit records)
        import_email_log(conn, wb_gas)

        # 10. Quotes (1 quote with pricing intelligence)
        import_quotes(conn, wb_gas)

        # 11. Intelligence benchmarks (11 facility types)
        import_intelligence_benchmarks(conn, wb_gas)

        # 12. Intelligence settings (14 pricing parameters)
        import_intelligence_settings(conn, wb_gas)

        # 13. Finance categories (14 expense categories)
        import_finance_categories(conn, wb_gas)

        log.info("")
        log.info("=" * 60)
        log.info("PHASE 2: Historical P&L Data (Template DB)")
        log.info("=" * 60)

        # 14. Historical invoices + expenses + snapshots (18 months × 6 sites)
        import_historical_pl(conn, wb_template)

        # 15. Inspections (12 records)
        import_template_inspections(conn, wb_template)

        # 16. Incidents (6 records)
        import_template_incidents(conn, wb_template)

        # 17. Contracts (6 records)
        import_template_contracts(conn, wb_template)

        log.info("")
        log.info("=" * 60)
        log.info("MIGRATION COMPLETE")
        log.info("=" * 60)

        # Summary — all tables
        for table in ["ops_cleaners", "pay_workers", "pay_entries",
                       "fin_invoices", "fin_expenses", "fin_transactions",
                       "fin_snapshots", "fin_settings",
                       "ops_inspections", "ops_incidents",
                       "entities", "opportunities", "opportunity_scores",
                       "contacts", "addresses", "entity_locations",
                       "entity_source_links", "outreach_activities",
                       "email_send_log",
                       "intelligence_benchmarks", "intelligence_settings",
                       "fin_expense_categories"]:
            try:
                count = db_pg.fetchval(conn, f"SELECT COUNT(*) FROM {table}")
                log.info("  %-25s %s rows", table, count)
            except Exception:
                log.info("  %-25s (table not found)", table)


if __name__ == "__main__":
    run_migration()
