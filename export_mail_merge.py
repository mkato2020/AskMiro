"""
export_mail_merge.py — AskMiro Lead Intelligence OS
Exports top scored leads to a formatted Excel file for mail merge / cold email outreach.

Usage:
    python export_mail_merge.py                          # top 20 leads, score >= 70
    python export_mail_merge.py --limit 10               # top 10
    python export_mail_merge.py --min-score 80           # higher bar
    python export_mail_merge.py --sector healthcare      # filter by sector
    python export_mail_merge.py --borough Lambeth        # filter by borough
"""

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))

import database
import outreach_generator

# ── Colour constants ───────────────────────────────────────────────────────────
NAVY_HEX        = "0D1C2E"
WHITE_HEX       = "FFFFFF"
GREEN_HEX       = "D6F0E0"   # score >= 80
YELLOW_HEX      = "FFF9C4"   # score >= 70
ORANGE_HEX      = "FFE0B2"   # score >= 60
TEAL_HEX        = "0A9688"

# ── Column definitions ─────────────────────────────────────────────────────────
COLUMNS = [
    ("A", "#",                    4),
    ("B", "Business Name",       30),
    ("C", "Sector",              18),
    ("D", "Borough",             18),
    ("E", "Decision Maker Role", 22),
    ("F", "Salutation",          28),
    ("G", "Phone",               16),
    ("H", "Website",             28),
    ("I", "Address",             35),
    ("J", "Score",                7),
    ("K", "HVT",                  6),
    ("L", "Recommended Offer",   35),
    ("M", "Buying Signals",      25),
    ("N", "Email Subject",       35),
    ("O", "Cold Email Body",     60),
    ("P", "Call Opener",         60),
    ("Q", "LinkedIn Intro",      60),
    ("R", "Follow Up Email",     60),
]

WRAP_COLS = {"O", "P", "Q", "R"}


# ── Data helpers ───────────────────────────────────────────────────────────────

def _fetch_leads(limit: int, min_score: int, sector=None, borough=None):
    conditions = ["priority_score >= ?"]
    params = [min_score]
    if sector:
        conditions.append("LOWER(normalized_sector) = LOWER(?)")
        params.append(sector)
    if borough:
        conditions.append("LOWER(borough) = LOWER(?)")
        params.append(borough)
    params.append(limit)

    with database.db_connection() as conn:
        rows = conn.execute(
            f"""SELECT * FROM lead_records
                WHERE {" AND ".join(conditions)}
                ORDER BY priority_score DESC, high_value_target DESC, review_count DESC
                LIMIT ?""",
            params,
        ).fetchall()
    return [dict(r) for r in rows]


def _extract_email_subject(cold_email: str, business_name: str) -> str:
    """Extract subject from cold_email or return a default."""
    if not cold_email:
        return f"Re: Cleaning services for {business_name}"
    for line in cold_email.splitlines():
        stripped = line.strip()
        if stripped.lower().startswith("subject:"):
            return stripped[len("subject:"):].strip()
    return f"Re: Cleaning services for {business_name}"


def _get_or_generate_outreach(row: dict) -> dict:
    """
    Returns outreach fields as a dict (keys: cold_email, call_opener,
    linkedin_intro, follow_up_email). Returns empty dict on failure.
    """
    place_id = row["place_id"]

    # Try to generate — returns None if already cached
    try:
        pkg = outreach_generator.generate_outreach(row, force=False)
    except Exception as exc:
        print(f"  [!] Outreach generation error for {row.get('business_name')}: {exc}")
        pkg = None

    if pkg is not None:
        # Freshly generated OutreachPackage dataclass
        return {
            "cold_email":     getattr(pkg, "cold_email",     "") or "",
            "call_opener":    getattr(pkg, "call_opener",    "") or "",
            "linkedin_intro": getattr(pkg, "linkedin_intro", "") or "",
            "follow_up_email":getattr(pkg, "follow_up_email","") or "",
        }

    # Already cached — fetch from DB
    cached = database.get_outreach(place_id)
    if cached:
        return {
            "cold_email":     cached.get("cold_email",     "") or "",
            "call_opener":    cached.get("call_opener",    "") or "",
            "linkedin_intro": cached.get("linkedin_intro", "") or "",
            "follow_up_email":cached.get("follow_up_email","") or "",
        }

    return {}


# ── Excel builder ──────────────────────────────────────────────────────────────

def _score_fill(score) -> PatternFill:
    if score is None:
        return None
    try:
        s = int(score)
    except (ValueError, TypeError):
        return None
    if s >= 80:
        return PatternFill("solid", fgColor=GREEN_HEX)
    if s >= 70:
        return PatternFill("solid", fgColor=YELLOW_HEX)
    if s >= 60:
        return PatternFill("solid", fgColor=ORANGE_HEX)
    return None


def _build_instructions_sheet(wb):
    ws = wb.create_sheet("How to Mail Merge")
    ws.column_dimensions["A"].width = 90

    # Row 1 — title
    ws["A1"] = "Mail Merge Instructions"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY_HEX)
    ws.row_dimensions[1].height = 28

    # Row 3 — Microsoft Word header
    ws["A3"] = "In Microsoft Word:"
    ws["A3"].font = Font(name="Arial", size=10, bold=True, color=NAVY_HEX)

    # Rows 4-8 — Word steps
    word_steps = [
        "1. Open Microsoft Word and create a new document with your email template.",
        "2. Go to Mailings > Start Mail Merge > Email Messages (or Letters).",
        "3. Click Select Recipients > Use an Existing List and choose this Excel file.",
        "4. Select the 'Leads' sheet when prompted.",
        "5. Insert merge fields using Mailings > Insert Merge Field — use the column headers listed below.",
    ]
    for i, step in enumerate(word_steps, start=4):
        ws.cell(row=i, column=1, value=step).font = Font(name="Arial", size=10)

    # Row 10 — Google Docs header
    ws["A10"] = "In Google Docs:"
    ws["A10"].font = Font(name="Arial", size=10, bold=True, color=NAVY_HEX)

    # Rows 11-14 — Google Docs steps
    gdocs_steps = [
        "1. Upload this Excel file to Google Drive and open it in Google Sheets.",
        "2. In Google Docs, open your email template document.",
        "3. Install the 'Mail Merge for Gmail' add-on (or use Google Workspace's built-in merge).",
        "4. Point the merge tool to the 'Leads' sheet and map columns using the field names below.",
    ]
    for i, step in enumerate(gdocs_steps, start=11):
        ws.cell(row=i, column=1, value=step).font = Font(name="Arial", size=10)

    # Row 16 — field names header
    ws["A16"] = "Field names to use in your template:"
    ws["A16"].font = Font(name="Arial", size=10, bold=True, color=NAVY_HEX)

    row = 17
    for _col_letter, header, _width in COLUMNS:
        field_name = header.replace(" ", "")  # e.g. "Business Name" -> "BusinessName"
        ws.cell(row=row, column=1, value=f"{{{{{field_name}}}}}  —  {header}")
        ws.cell(row=row, column=1).font = Font(name="Arial", size=9, color="333333")
        row += 1


def build_excel(leads: list, outreach_map: dict, out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=NAVY_HEX)
    header_font = Font(name="Arial", size=10, bold=True, color=WHITE_HEX)

    for col_idx, (col_letter, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28

    # Freeze top row and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{COLUMNS[-1][0]}1"

    # ── Data rows ─────────────────────────────────────────────────────────────
    teal_font  = Font(name="Arial", size=9, color=TEAL_HEX, bold=True)
    plain_font = Font(name="Arial", size=9)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    top_align  = Alignment(vertical="top", wrap_text=False)

    for row_num, lead in enumerate(leads, start=2):
        place_id = lead.get("place_id", "")
        outreach = outreach_map.get(place_id, {})

        dm_role  = lead.get("ai_decision_maker_type") or "Facilities Manager"
        sector   = (lead.get("normalized_sector") or "other").replace("_", " ").title()
        score    = lead.get("priority_score") or 0
        hvt      = lead.get("high_value_target")
        cold_email = outreach.get("cold_email", "")

        row_data = [
            row_num - 1,                              # A: #
            lead.get("business_name") or "",          # B: Business Name
            sector,                                   # C: Sector
            lead.get("borough") or "",                # D: Borough
            dm_role,                                  # E: Decision Maker Role
            "Dear " + dm_role,                        # F: Salutation
            lead.get("phone") or "",                  # G: Phone
            lead.get("website") or "",                # H: Website
            lead.get("address") or "",                # I: Address
            score,                                    # J: Score
            "★ Yes" if hvt else "",                  # K: HVT
            lead.get("recommended_offer") or "",      # L: Recommended Offer
            lead.get("trigger_summary") or "",        # M: Buying Signals
            _extract_email_subject(cold_email, lead.get("business_name", "")),  # N: Email Subject
            cold_email,                               # O: Cold Email Body
            outreach.get("call_opener", ""),          # P: Call Opener
            outreach.get("linkedin_intro", ""),       # Q: LinkedIn Intro
            outreach.get("follow_up_email", ""),      # R: Follow Up Email
        ]

        for col_idx, value in enumerate(row_data, start=1):
            col_letter = COLUMNS[col_idx - 1][0]
            cell = ws.cell(row=row_num, column=col_idx, value=value)

            if col_letter in WRAP_COLS:
                cell.font = Font(name="Arial", size=9)
                cell.alignment = wrap_align
            else:
                cell.font = plain_font
                cell.alignment = top_align

            # Score colour-coding (column J)
            if col_letter == "J":
                fill = _score_fill(score)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=9, bold=True)

            # HVT teal text (column K)
            if col_letter == "K" and value:
                cell.font = teal_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Instructions sheet ───────────────────────────────────────────────────
    _build_instructions_sheet(wb)

    wb.save(str(out_path))


# ── Core logic (shared by CLI and API) ────────────────────────────────────────

def run_mail_merge_export(
    limit: int = 20,
    min_score: int = 70,
    sector: str = None,
    borough: str = None,
    out_path: Path = None,
) -> tuple:
    """
    Fetch leads, generate/fetch outreach, build Excel.
    Returns (out_path, leads_exported, outreach_generated).
    """
    leads = _fetch_leads(limit, min_score, sector, borough)
    if not leads:
        return None, 0, 0

    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = Path.home() / "Downloads" / f"askmiro_mail_merge_{ts}.xlsx"

    outreach_map = {}
    outreach_generated = 0
    total = len(leads)

    print(f"Processing outreach for {total} leads...")

    for i, lead in enumerate(leads, start=1):
        place_id = lead["place_id"]
        biz_name = lead.get("business_name", place_id)
        print(f"  [{i}/{total}] {biz_name}", end="", flush=True)

        # Check if outreach already exists to avoid counting cached ones as generated
        existing = database.get_outreach(place_id)
        if existing:
            print(" (cached)")
            outreach_map[place_id] = {
                "cold_email":      existing.get("cold_email",     "") or "",
                "call_opener":     existing.get("call_opener",    "") or "",
                "linkedin_intro":  existing.get("linkedin_intro", "") or "",
                "follow_up_email": existing.get("follow_up_email","") or "",
            }
        else:
            print(" (generating...)", end="", flush=True)
            result = _get_or_generate_outreach(lead)
            if result:
                outreach_generated += 1
                print(" done")
            else:
                print(" failed")
            outreach_map[place_id] = result
            # Rate limiting: stay under 50 RPM
            if i < total:
                time.sleep(1.3)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_excel(leads, outreach_map, out_path)
    return out_path, len(leads), outreach_generated


# ── Campaign export helper (accepts specific place_ids) ───────────────────────

def build_mail_merge_excel(place_ids: list, out_path: Path) -> Path:
    """
    Build a mail merge Excel file for a specific list of place_ids.
    Used by the campaign export endpoint.
    Returns the output path.
    """
    if not place_ids:
        raise ValueError("place_ids list is empty")

    # Fetch the specific leads by place_id
    placeholders = ",".join("?" * len(place_ids))
    with database.db_connection() as conn:
        rows = conn.execute(
            f"SELECT * FROM lead_records WHERE place_id IN ({placeholders})",
            place_ids,
        ).fetchall()
    leads = [dict(r) for r in rows]

    if not leads:
        raise ValueError("No lead records found for the given place_ids")

    # Build outreach map
    outreach_map = {}
    for lead in leads:
        place_id = lead["place_id"]
        cached = database.get_outreach(place_id)
        if cached:
            outreach_map[place_id] = {
                "cold_email":      cached.get("cold_email",     "") or "",
                "call_opener":     cached.get("call_opener",    "") or "",
                "linkedin_intro":  cached.get("linkedin_intro", "") or "",
                "follow_up_email": cached.get("follow_up_email","") or "",
            }
        else:
            result = _get_or_generate_outreach(lead)
            outreach_map[place_id] = result

    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_excel(leads, outreach_map, out_path)
    return out_path


# ── CLI entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export top leads to an Excel mail merge file"
    )
    parser.add_argument("--limit",     type=int, default=20,
                        help="Number of leads to export (default 20)")
    parser.add_argument("--min-score", type=int, default=70,
                        help="Minimum priority score (default 70)")
    parser.add_argument("--sector",    type=str, default=None,
                        help="Filter by sector (e.g. healthcare)")
    parser.add_argument("--borough",   type=str, default=None,
                        help="Filter by borough (e.g. Lambeth)")
    parser.add_argument("--out",       type=str, default=None,
                        help="Custom output path (.xlsx)")
    args = parser.parse_args()

    out_path = Path(args.out) if args.out else None

    out_file, exported, generated = run_mail_merge_export(
        limit     = args.limit,
        min_score = args.min_score,
        sector    = args.sector,
        borough   = args.borough,
        out_path  = out_path,
    )

    if out_file is None:
        print("No leads found matching those filters.")
        sys.exit(1)

    print()
    print(f"Excel exported: {out_file}")
    print(f"  Leads exported:           {exported}")
    print(f"  AI outreach generated:    {generated}")
    print(f"  Outreach from cache:      {exported - generated}")


if __name__ == "__main__":
    main()
